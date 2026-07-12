from __future__ import annotations

import argparse
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SHEET_TABLES = {
    "Propietarios": (
        "crm_owners",
        {
            "id": "id",
            "createdAt": "created_at",
            "name": "name",
            "phone": "phone",
            "email": "email",
            "notes": "notes",
        },
    ),
    "Clientes": (
        "crm_clients",
        {
            "id": "id",
            "createdAt": "created_at",
            "name": "name",
            "phone": "phone",
            "email": "email",
            "interest": "interest",
            "notes": "notes",
        },
    ),
    "Inmuebles": (
        "crm_properties",
        {
            "id": "id",
            "createdAt": "created_at",
            "title": "title",
            "address": "address",
            "price": "price",
            "status": "status",
            "details": "details",
            "ownerId": "owner_id",
        },
    ),
    "Actividades": (
        "crm_activities",
        {
            "id": "id",
            "createdAt": "created_at",
            "clientId": "client_id",
            "propertyId": "property_id",
            "type": "type",
            "date": "activity_date",
            "notes": "notes",
            "ownerId": "owner_id",
        },
    ),
    "Tareas": (
        "crm_tasks",
        {
            "id": "id",
            "createdAt": "created_at",
            "title": "title",
            "dueDate": "due_date",
            "priority": "priority",
            "clientId": "client_id",
            "propertyId": "property_id",
            "notes": "notes",
            "done": "done",
            "ownerId": "owner_id",
        },
    ),
}

SHEET_ORDER = ["Propietarios", "Clientes", "Inmuebles", "Actividades", "Tareas"]
TEXT_COLUMNS = {"id", "name", "phone", "email", "notes", "interest", "title", "address", "status", "details", "owner_id", "client_id", "property_id", "type", "priority"}
DATE_COLUMNS = {"activity_date", "due_date"}
TIMESTAMP_COLUMNS = {"created_at"}
NUMERIC_COLUMNS = {"price"}
BOOLEAN_COLUMNS = {"done"}
FK_COLUMNS = {
    "client_id": "Clientes",
    "owner_id": "Propietarios",
    "property_id": "Inmuebles",
}


def normalize(value: Any, column: str) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    if column in TEXT_COLUMNS:
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)
    return value


def sql_literal(value: Any, column: str) -> str:
    value = normalize(value, column)
    if value is None:
        return "NULL"
    if column in BOOLEAN_COLUMNS:
        return "true" if bool(value) else "false"
    if column in NUMERIC_COLUMNS:
        return str(value)
    if column in DATE_COLUMNS:
        if isinstance(value, datetime):
            value = value.date()
        if isinstance(value, date):
            return f"'{value.isoformat()}'"
    if column in TIMESTAMP_COLUMNS:
        if isinstance(value, datetime):
            return f"'{value.isoformat()}'"
    text = str(value).replace("'", "''")
    return f"'{text}'"


def collect_ids(workbook: Any) -> dict[str, set[str]]:
    ids: dict[str, set[str]] = {}

    for sheet_name in ["Propietarios", "Clientes", "Inmuebles"]:
        worksheet = workbook[sheet_name]
        headers = [cell.value for cell in worksheet[1]]
        id_index = headers.index("id")
        ids[sheet_name] = {
            str(row[id_index]).strip()
            for row in worksheet.iter_rows(min_row=2, values_only=True)
            if row[id_index] is not None
        }

    return ids


def rows_for_sheet(workbook: Any, sheet_name: str, ids: dict[str, set[str]]) -> tuple[str, list[str], list[list[Any]], list[str]]:
    table, mapping = SHEET_TABLES[sheet_name]
    worksheet = workbook[sheet_name]
    headers = [cell.value for cell in worksheet[1]]
    columns = [mapping[header] for header in headers if header in mapping]
    rows: list[list[Any]] = []
    warnings: list[str] = []

    for raw_row in worksheet.iter_rows(min_row=2, values_only=True):
        source = dict(zip(headers, raw_row))
        if not any(value is not None for value in source.values()):
            continue
        row = []
        row_id = source.get("id") or "sin_id"

        for header in headers:
            if header not in mapping:
                continue
            column = mapping[header]
            value = source.get(header)
            target_sheet = FK_COLUMNS.get(column)

            if target_sheet and value is not None and str(value).strip() not in ids[target_sheet]:
                warnings.append(f"-- Aviso: {sheet_name} {row_id} tenia {column}={value}, pero no existe en {target_sheet}. Se importa como NULL.")
                value = None

            row.append(value)

        rows.append(row)

    return table, columns, rows, warnings


def upsert_statement(table: str, columns: list[str], rows: list[list[Any]]) -> str:
    if not rows:
        return f"-- {table}: sin registros para importar\n"

    values = []
    for row in rows:
        values.append("(" + ", ".join(sql_literal(value, column) for value, column in zip(row, columns)) + ")")

    update_columns = [column for column in columns if column != "id"]
    updates = ", ".join(f"{column} = excluded.{column}" for column in update_columns)

    return (
        f"insert into {table} ({', '.join(columns)})\n"
        "values\n  "
        + ",\n  ".join(values)
        + f"\non conflict (id) do update set {updates};\n"
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()

    workbook = load_workbook(args.source, data_only=True)
    blocks = [
        "-- Importacion de datos legacy CRM Aktua.",
        "-- Seguro para reejecutar: usa upsert por id.",
        "begin;",
    ]
    ids = collect_ids(workbook)

    for sheet_name in SHEET_ORDER:
        table, columns, rows, warnings = rows_for_sheet(workbook, sheet_name, ids)
        blocks.append(f"\n-- {sheet_name}: {len(rows)} registros")
        blocks.extend(warnings)
        blocks.append(upsert_statement(table, columns, rows))

    blocks.append("commit;\n")
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text("\n".join(blocks), encoding="utf-8")
    print(args.output)


if __name__ == "__main__":
    main()
